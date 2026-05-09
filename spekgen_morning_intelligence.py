#!/usr/bin/env python3
"""
spekgen_morning_intelligence.py
SPEKGEN Agency — Morning Intelligence Runner v2.0
==================================================
Corre diario a las 8am (México):
  1. Descarga AD_LOG desde Google Drive (evita FUSE deadlock)
  2. Descarga .env de cada cliente desde Drive para obtener META_ACCESS_TOKEN
  3. Jala métricas de Meta (últimos 7 días) para LO FITNESS, GREENRAY, HC
  4. Detecta ads con CPA > threshold → PAUSA automáticamente + actualiza AD LOG
  5. Detecta ads con >= N días activos → alerta de rotación creativa
  6. Jala tareas abiertas de ClickUp por cliente
  7. Genera Narrative Memo (SITUATION / COMPLICATIONS / RESOLUTION)
  8. Guarda memo como Google Doc en Drive

USO:
  python3 spekgen_morning_intelligence.py             # Producción
  python3 spekgen_morning_intelligence.py --dry-run   # Sin cambios reales
  python3 spekgen_morning_intelligence.py --client LO_FITNESS

REQUIERE (.env local en raíz del repo, NUNCA commitear):
  GOOGLE_ACCESS_TOKEN=ya29.xxx    # OAuth token con permisos Drive
  CLICKUP_TOKEN=pk_xxx            # ClickUp personal token

REQUIERE (pip):
  pip install requests openpyxl python-dotenv google-auth google-auth-httplib2 google-api-python-client
"""

import os
import sys
import json
import argparse
import tempfile
import requests
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────

META_API_VERSION = "v21.0"
META_BASE_URL    = f"https://graph.facebook.com/{META_API_VERSION}"
DRIVE_API_BASE   = "https://www.googleapis.com/drive/v3"
DRIVE_UPLOAD_URL = "https://www.googleapis.com/upload/drive/v3"

# Drive IDs — actualizar cuando se confirmen con Gibran
AD_LOG_UNIFIED_DRIVE_ID = "PLACEHOLDER_SPEKGEN_CLIENTS_AD_LOG_DRIVE_ID"  # SPEKGEN_CLIENTS_AD_LOG_v2.0.xlsx
REPORTS_FOLDER_DRIVE_ID = "PLACEHOLDER_REPORTS_FOLDER_DRIVE_ID"           # SPK-10.AUTOMATION/morning-intelligence/reports/

# Learning phase (Meta sale de aprendizaje con ~50 compras en 7 días)
LEARNING_PURCHASE_THRESHOLD = 50
LEARNING_CPA_MULTIPLIER     = 1.30  # +30% de tolerancia durante aprendizaje

# ─────────────────────────────────────────────────────────────────────────────
#  CPA THRESHOLDS
# ─────────────────────────────────────────────────────────────────────────────

LF_CPA_THRESHOLDS = {
    "fit max":         398,   # $598 - 200
    "metafit":         298,   # $498 - 200
    "meta fit":        298,
    "ph detox":        289,   # $489 - 200
    "omega3":          289,
    "omega 3":         289,
    "oceane":          289,
    "potasio":         189,   # $389 - 200
    "magnesio":        189,
    "6r mask":          68,   # $268 - 200
    "shampoo":          48,   # $248 - 200
    "acondicionador":   48,
    "aquacell":        None,  # $128 — solo alerta, no auto-pause
    "aqua cell":       None,
    "shinewave":       None,  # $178
    "forcetonic":      None,  # $198
    "fitbar":          None,  # $98
    "default":         289,
}

GR_CPA_THRESHOLDS = {
    "gaxaliv":              None,   # $158 — margen negativo para ads
    "gax-aliv":             None,
    "artrix":                95,    # $295 - 200
    "motiox":               None,   # $189 — por debajo del mínimo
    "h2fx micfend":         259,    # $459 - 200
    "micfend":              259,
    "h2fx c-ynergy":        259,
    "c-ynergy":             259,
    "micbio artemyr":       192,    # $392 - 200
    "artemyr":              192,
    "carmor 4":              40,    # $240 - 200
    "carmor 5":              76,    # $276 - 200
    "h2fx fluxart":         259,
    "fluxart":              259,
    "h2fx synap":           259,
    "synap":                259,
    "h2fx balanz":          259,
    "balanz":               259,
    "intrab sleep":         172,    # $372 - 200
    "intrab":               172,
    "bellsan ultra":        210,    # $410 - 200
    "bellsan":              210,
    "detofx":               161,    # $361 - 200
    "benfotx":              203,    # $403 - 200
    "metaplex ayuno":       259,    # $459 - 200
    "metaplex":             192,    # $392 fallback
    "h2fx meta":            259,
    "hormofx m 40+":        210,    # $410 - 200
    "hormofx reset":        265,    # $465 - 200
    "hormofx men 40+":      265,
    "hormofx":              161,    # $361 fallback
    "hidronex cell":        469,    # $669 - 200
    "agua de mar hi galón": 259,    # $459 - 200
    "agua de mar hi":       None,   # formatos 1lt/$104, 500ml/$55
    "vitaminas+":           128,    # $328 - 200
    "vitaminas":            128,
    "g-xamin":              210,    # $410 - 200
    "glucosamina+":          95,    # $295 - 200
    "colágeno fitness":     712,    # $912 - 200
    "colágeno nutrition":   712,
    "colágeno beauty":      712,
    "colageno":             712,    # fallback
    "orgon protein complex 40+akg": 1126,   # $1326 - 200
    "orgon protein complex 40+tmg": 1126,
    "orgon protein complex active":  904,   # $1104 - 200
    "orgon protein complex 40+":     904,
    "orgon protein":                 904,   # fallback
    "creatine muscle complex":       551,   # $751 - 200
    "creatine":                      551,
    "default":                       192,   # threshold mínimo viable ($392 avg)
}

# ─────────────────────────────────────────────────────────────────────────────
#  CLIENTS
# ─────────────────────────────────────────────────────────────────────────────

CLIENTS = {
    "LO_FITNESS": {
        "name":            "LO FITNESS",
        "env_drive_id":    "PLACEHOLDER_LF_ENV_DRIVE_ID",      # LF - 10. LOGS/02. META API SCRIPTS/.env
        "ad_log_drive_id": "1MSY4SQDu4W5lyUR9ncghvOKV8WRwuSmI",# LO_FITNESS_AD_LOG_v1.0.xlsx
        "ad_log_sheet":    "\U0001f4cb LO FITNESS ADS",
        "ad_prefix":       "LF",
        "clickup_list_id": "901711952297",
        "cpa_thresholds":  LF_CPA_THRESHOLDS,
        "rotation_days":   7,
    },
    "GREENRAY": {
        "name":            "GREENRAY",
        "env_drive_id":    "PLACEHOLDER_GR_ENV_DRIVE_ID",       # GR - GREENRAY/.env
        "ad_log_drive_id": AD_LOG_UNIFIED_DRIVE_ID,             # usa log unificado
        "ad_log_sheet":    "\U0001f4cb GREENRAY ADS",
        "ad_prefix":       "GR",
        "clickup_list_id": "901711854083",
        "cpa_thresholds":  GR_CPA_THRESHOLDS,
        "rotation_days":   7,
    },
    "HEALTHY_CHUCHOS": {
        "name":            "HEALTHY CHUCHOS",
        "env_drive_id":    "PLACEHOLDER_HC_ENV_DRIVE_ID",       # HC - HEALTHY CHUCHOS/.env
        "ad_log_drive_id": AD_LOG_UNIFIED_DRIVE_ID,
        "ad_log_sheet":    "\U0001f4cb HC ADS",
        "ad_prefix":       "HC",
        "clickup_list_id": "PLACEHOLDER_HC_CLICKUP_LIST_ID",
        "cpa_thresholds":  {"default": 150},
        "rotation_days":   7,
    },
    "F24_FERRETERIA": {
        "name":            "F24 FERRETERIA",
        "env_drive_id":    None,   # sin Meta Ads — sprint Shopify
        "ad_log_drive_id": None,
        "ad_log_sheet":    None,
        "ad_prefix":       "F24",
        "clickup_list_id": "901713618377",  # Fase 1 (actualizar a 901713618380 en Fase 2, 901713618382 en Fase 3)
        "cpa_thresholds":  {},
        "rotation_days":   7,
    },
}

# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE DRIVE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_google_token() -> str:
    """Obtiene el access token de Google desde env o service account."""
    token = os.environ.get("GOOGLE_ACCESS_TOKEN")
    if token:
        return token
    try:
        import google.auth
        import google.auth.transport.requests
        credentials, _ = google.auth.default(
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        credentials.refresh(google.auth.transport.requests.Request())
        return credentials.token
    except Exception:
        pass
    raise RuntimeError(
        "No se encontró GOOGLE_ACCESS_TOKEN en env ni credenciales de service account.\n"
        "Agrega GOOGLE_ACCESS_TOKEN=ya29.xxx al .env local (no commitear)."
    )


def download_drive_file(file_id: str, dest_path: Path, token: str) -> Path:
    """Descarga un archivo de Drive a disco local — evita FUSE deadlock."""
    url = f"{DRIVE_API_BASE}/files/{file_id}?alt=media"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    dest_path.write_bytes(r.content)
    return dest_path


def download_drive_text(file_id: str, token: str) -> str:
    """Descarga un archivo de texto de Drive como string."""
    tmp = Path(tempfile.gettempdir()) / f"spekgen_tmp_{file_id}.txt"
    download_drive_file(file_id, tmp, token)
    content = tmp.read_text(encoding="utf-8")
    tmp.unlink(missing_ok=True)
    return content


def upload_drive_file(file_id: str, local_path: Path, token: str) -> bool:
    """Re-sube un archivo modificado a Drive (actualiza en-place)."""
    url = f"{DRIVE_UPLOAD_URL}/files/{file_id}?uploadType=media"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    with open(local_path, "rb") as f:
        r = requests.patch(
            url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": mime},
            data=f,
            timeout=120,
        )
    return r.status_code == 200


def create_google_doc(title: str, content: str, parent_folder_id: str, token: str) -> str | None:
    """Crea un Google Doc con el contenido. Devuelve el fileId o None."""
    metadata = {
        "name": title,
        "mimeType": "application/vnd.google-apps.document",
        "parents": [parent_folder_id],
    }
    r = requests.post(
        f"{DRIVE_API_BASE}/files",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=metadata,
        timeout=30,
    )
    if r.status_code not in (200, 201):
        print(f"   ✗ Error creando Google Doc: {r.text[:200]}")
        return None
    file_id = r.json().get("id")
    upload_url = f"{DRIVE_UPLOAD_URL}/files/{file_id}?uploadType=media"
    r2 = requests.patch(
        upload_url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "text/plain; charset=utf-8"},
        data=content.encode("utf-8"),
        timeout=60,
    )
    return file_id if r2.status_code == 200 else None


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS GENÉRICOS
# ─────────────────────────────────────────────────────────────────────────────

def col_letter_to_index(letter: str) -> int:
    idx = 0
    for ch in letter.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def get_cell(row, col_letter: str):
    idx = col_letter_to_index(col_letter)
    return row[idx - 1].value if idx <= len(row) else None


def load_env_from_text(text: str) -> dict:
    env = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        env[key.strip()] = val.strip().strip('"').strip("'")
    return env


def load_env_from_drive(drive_id: str, token: str) -> dict:
    """Descarga y parsea un .env desde Drive."""
    try:
        text = download_drive_text(drive_id, token)
        return load_env_from_text(text)
    except Exception as e:
        print(f"   ✗ No se pudo cargar .env (Drive ID: {drive_id}): {e}")
        return {}


# ─────────────────────────────────────────────────────────────────────────────
#  META API
# ─────────────────────────────────────────────────────────────────────────────

def meta_get(endpoint: str, token: str, params: dict = None) -> dict:
    url = f"{META_BASE_URL}/{endpoint}"
    p = {**(params or {}), "access_token": token}
    r = requests.get(url, params=p, timeout=30)
    result = r.json()
    if "error" in result:
        err = result["error"]
        raise RuntimeError(f"[{err.get('code')}] {err.get('type')}: {err.get('message')}")
    return result


def meta_post(endpoint: str, token: str, data: dict = None) -> dict:
    url = f"{META_BASE_URL}/{endpoint}"
    d = {**(data or {}), "access_token": token}
    r = requests.post(url, data=d, timeout=30)
    result = r.json()
    if "error" in result:
        err = result["error"]
        raise RuntimeError(f"[{err.get('code')}] {err.get('type')}: {err.get('message')}")
    return result


def get_ad_insights(meta_ad_id: str, token: str, days: int = 7) -> dict:
    """Jala métricas del ad (últimos N días). Devuelve spend, purchases, CPA, CTR."""
    try:
        result = meta_get(
            f"{meta_ad_id}/insights",
            token,
            params={
                "date_preset": f"last_{days}d",
                "fields": "spend,actions,cost_per_action_type,impressions,clicks,ctr",
                "level": "ad",
            },
        )
        data = result.get("data", [])
        if not data:
            return {"spend": 0.0, "purchases": 0, "cpa": None, "impressions": 0, "clicks": 0, "ctr": 0.0, "days": days}

        d = data[0]
        spend       = float(d.get("spend", 0) or 0)
        impressions = int(d.get("impressions", 0) or 0)
        clicks      = int(d.get("clicks", 0) or 0)
        ctr         = float(d.get("ctr", 0) or 0)

        purchases = 0
        for action in d.get("actions", []):
            if action.get("action_type") in [
                "purchase",
                "offsite_conversion.fb_pixel_purchase",
                "omni_purchase",
            ]:
                purchases += int(float(action.get("value", 0)))

        cpa = round(spend / purchases, 2) if purchases > 0 else None
        return {
            "spend": round(spend, 2),
            "purchases": purchases,
            "cpa": cpa,
            "impressions": impressions,
            "clicks": clicks,
            "ctr": round(ctr, 4),
            "days": days,
        }
    except Exception as e:
        return {"error": str(e), "spend": 0.0, "purchases": 0, "cpa": None, "days": days}


def pause_ad_in_meta(meta_ad_id: str, token: str, dry_run: bool = False) -> bool:
    if dry_run:
        print(f"         [DRY RUN] Pausaría meta_ad_id: {meta_ad_id}")
        return True
    try:
        result = meta_post(meta_ad_id, token, {"status": "PAUSED"})
        return bool(result.get("success"))
    except Exception as e:
        print(f"         ✗ Error al pausar {meta_ad_id}: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
#  AD LOG
# ─────────────────────────────────────────────────────────────────────────────

def get_active_ads(log_path: Path, sheet_name: str, prefix: str) -> list:
    try:
        wb = openpyxl.load_workbook(str(log_path), data_only=True)
        ws = wb[sheet_name]
        ads = []
        for row in ws.iter_rows(min_row=4):
            row_id  = get_cell(row, "A")
            status  = get_cell(row, "W")
            meta_id = get_cell(row, "AT")
            if not row_id or not str(row_id).startswith(prefix):
                continue
            if str(status).strip() != "Activo":
                continue
            if not meta_id:
                continue
            ads.append({
                "ad_id":           str(row_id).strip(),
                "meta_ad_id":      str(meta_id).strip(),
                "product":         str(get_cell(row, "K") or "").strip().lower(),
                "activation_date": get_cell(row, "E"),
            })
        return ads
    except Exception as e:
        print(f"   ✗ Error leyendo AD LOG ({sheet_name}): {e}")
        return []


def get_pending_ads(log_path: Path, sheet_name: str, prefix: str) -> list:
    try:
        wb = openpyxl.load_workbook(str(log_path), data_only=True)
        ws = wb[sheet_name]
        pending = []
        for row in ws.iter_rows(min_row=4):
            row_id = get_cell(row, "A")
            status = get_cell(row, "W")
            if not row_id or not str(row_id).startswith(prefix):
                continue
            if str(status).strip() in ["Draft", "En Producción"]:
                pending.append({
                    "ad_id":   str(row_id).strip(),
                    "product": str(get_cell(row, "K") or "").strip(),
                    "status":  str(status).strip(),
                })
        return pending
    except Exception:
        return []


def update_log_status(log_path: Path, sheet_name: str, ad_id: str, new_status: str):
    try:
        wb = openpyxl.load_workbook(str(log_path))
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4):
            if str(get_cell(row, "A") or "").strip() == ad_id:
                ws.cell(row[0].row, col_letter_to_index("W")).value = new_status
                break
        wb.save(str(log_path))
    except Exception as e:
        print(f"   ✗ Error actualizando LOG para {ad_id}: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  CLICKUP
# ─────────────────────────────────────────────────────────────────────────────

def get_clickup_open_tasks(list_id: str, token: str) -> list:
    """Jala tareas abiertas/en progreso de una lista de ClickUp."""
    if not token or not list_id or "PLACEHOLDER" in list_id:
        return []
    try:
        url = f"https://api.clickup.com/api/v2/list/{list_id}/task"
        r = requests.get(
            url,
            headers={"Authorization": token},
            params={
                "statuses[]": ["open", "in progress"],
                "include_closed": False,
                "subtasks": True,
            },
            timeout=30,
        )
        tasks = r.json().get("tasks", [])
        now_ms = datetime.now().timestamp() * 1000
        result = []
        for t in tasks:
            due = t.get("due_date")
            overdue = bool(due and int(due) < now_ms)
            result.append({
                "id":       t.get("id"),
                "name":     t.get("name"),
                "status":   t.get("status", {}).get("status", ""),
                "priority": (t.get("priority") or {}).get("priority", ""),
                "due_date": due,
                "overdue":  overdue,
            })
        return result
    except Exception as e:
        print(f"   ✗ Error jalando ClickUp tasks ({list_id}): {e}")
        return []


# ─────────────────────────────────────────────────────────────────────────────
#  LÓGICA POR CLIENTE
# ─────────────────────────────────────────────────────────────────────────────

def resolve_cpa_threshold(product: str, thresholds: dict) -> int | None:
    product_lower = product.lower()
    for key, val in thresholds.items():
        if key != "default" and key in product_lower:
            return val
    return thresholds.get("default")


def parse_activation_date(raw) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(raw).strip(), fmt)
        except ValueError:
            pass
    return None


def run_client(
    client_key: str,
    config: dict,
    google_token: str,
    clickup_token: str,
    dry_run: bool,
) -> dict:
    print(f"\n{'─'*60}")
    print(f"  \U0001f4cc  {config['name']}")
    print(f"{'─'*60}")

    report = {
        "client":          config["name"],
        "client_key":      client_key,
        "clickup_list_id": config["clickup_list_id"],
        "timestamp":       datetime.now().isoformat(),
        "ads_paused":      [],
        "rotation_alerts": [],
        "draft_queue":     [],
        "clickup_tasks":   {"open": 0, "overdue": 0, "tasks": []},
        "errors":          [],
        "summary":         {},
    }

    # ── ClickUp tasks ────────────────────────────────────────────────────────
    cu_tasks = get_clickup_open_tasks(config["clickup_list_id"], clickup_token)
    overdue  = [t for t in cu_tasks if t["overdue"]]
    report["clickup_tasks"] = {"open": len(cu_tasks), "overdue": len(overdue), "tasks": cu_tasks}
    print(f"   \U0001f4cb  ClickUp: {len(cu_tasks)} tareas abiertas | {len(overdue)} vencidas")

    # ── Clientes sin Meta Ads (F24 en sprint Shopify) ────────────────────────
    if not config.get("env_drive_id") or not config.get("ad_log_drive_id"):
        print(f"   ℹ️  Sin Meta Ads configurados — solo ClickUp")
        report["summary"] = {
            "total_active": 0, "paused_today": 0, "rotation_alerts": 0,
            "drafts": 0, "in_production": 0, "errors": 0,
        }
        return report

    # ── Cargar .env desde Drive ───────────────────────────────────────────────
    env = load_env_from_drive(config["env_drive_id"], google_token)
    meta_token = (
        env.get("META_TOKEN") or
        env.get("META_ACCESS_TOKEN") or
        env.get("ACCESS_TOKEN") or
        env.get("meta_access_token")
    )
    if not meta_token:
        msg = f".env sin META_ACCESS_TOKEN (Drive ID: {config['env_drive_id']})"
        print(f"   ✗ {msg}")
        report["errors"].append(msg)
        report["summary"] = {
            "total_active": 0, "paused_today": 0, "rotation_alerts": 0,
            "drafts": 0, "in_production": 0, "errors": 1,
        }
        return report
    print(f"   ✔  Token Meta cargado")

    # ── Descargar AD LOG desde Drive ─────────────────────────────────────────
    tmp_dir  = Path(tempfile.gettempdir())
    log_path = tmp_dir / f"spekgen_ad_log_{client_key}.xlsx"
    try:
        download_drive_file(config["ad_log_drive_id"], log_path, google_token)
        print(f"   ✔  AD LOG descargado")
    except Exception as e:
        msg = f"No se pudo descargar AD LOG (Drive ID: {config['ad_log_drive_id']}): {e}"
        print(f"   ✗ {msg}")
        report["errors"].append(msg)
        report["summary"] = {
            "total_active": 0, "paused_today": 0, "rotation_alerts": 0,
            "drafts": 0, "in_production": 0, "errors": 1,
        }
        return report

    active_ads  = get_active_ads(log_path, config["ad_log_sheet"], config["ad_prefix"])
    pending_ads = get_pending_ads(log_path, config["ad_log_sheet"], config["ad_prefix"])
    report["draft_queue"] = pending_ads
    print(f"   \U0001f4ca  Activos: {len(active_ads)} | Pendientes: {len(pending_ads)}")

    log_modified = False

    for ad in active_ads:
        print(f"\n   \U0001f4ce  {ad['ad_id']} ({ad['product'] or '—'})")
        insights = get_ad_insights(ad["meta_ad_id"], meta_token, days=7)

        if "error" in insights:
            msg = f"{ad['ad_id']}: {insights['error']}"
            print(f"     ⚠️  {msg}")
            report["errors"].append(msg)
            continue

        cpa_str = f"${insights['cpa']}" if insights["cpa"] else "sin compras"
        print(f"     Spend 7d: ${insights['spend']} | Compras: {insights['purchases']} | CPA: {cpa_str}")

        # CHECK 1: CPA alto
        max_cpa = resolve_cpa_threshold(ad["product"], config["cpa_thresholds"])
        if insights["cpa"] and max_cpa is not None:
            purchases_7d     = insights["purchases"]
            in_learning      = purchases_7d < LEARNING_PURCHASE_THRESHOLD
            eff_threshold    = (
                round(max_cpa * LEARNING_CPA_MULTIPLIER, 2) if in_learning else max_cpa
            )
            if in_learning:
                print(f"     \U0001f4da  Aprendizaje ({purchases_7d} < {LEARNING_PURCHASE_THRESHOLD}) → threshold ${max_cpa} → ${eff_threshold}")

            if insights["cpa"] > eff_threshold:
                label = "CPA alto (aprendizaje)" if in_learning else "CPA alto"
                print(f"     \U0001f6a8  CPA ${insights['cpa']} > ${eff_threshold} → PAUSANDO [{label}]")
                paused = pause_ad_in_meta(ad["meta_ad_id"], meta_token, dry_run)
                if paused or dry_run:
                    if not dry_run:
                        update_log_status(log_path, config["ad_log_sheet"], ad["ad_id"], f"Pausado — {label}")
                        log_modified = True
                    report["ads_paused"].append({
                        "ad_id":          ad["ad_id"],
                        "meta_ad_id":     ad["meta_ad_id"],
                        "product":        ad["product"],
                        "cpa":            insights["cpa"],
                        "threshold":      eff_threshold,
                        "threshold_base": max_cpa,
                        "in_learning":    in_learning,
                        "spend_7d":       insights["spend"],
                        "purchases":      insights["purchases"],
                    })

        # CHECK 2: Rotación creativa
        activation_dt = parse_activation_date(ad["activation_date"])
        if activation_dt:
            days_active = (datetime.now() - activation_dt).days
            if days_active >= config["rotation_days"]:
                print(f"     \U0001f504  {days_active} días activo → alerta rotación")
                report["rotation_alerts"].append({
                    "ad_id":       ad["ad_id"],
                    "product":     ad["product"],
                    "days_active": days_active,
                    "spend_7d":    insights["spend"],
                    "purchases":   insights["purchases"],
                    "cpa":         insights["cpa"],
                })

    # Re-subir AD LOG si fue modificado
    if log_modified and not dry_run:
        print(f"\n   ↑ Re-subiendo AD LOG a Drive...")
        ok = upload_drive_file(config["ad_log_drive_id"], log_path, google_token)
        print(f"   {'✔' if ok else '✗'}  Upload {'exitoso' if ok else 'fallido'}")

    log_path.unlink(missing_ok=True)

    report["summary"] = {
        "total_active":    len(active_ads),
        "paused_today":    len(report["ads_paused"]),
        "rotation_alerts": len(report["rotation_alerts"]),
        "drafts":          len([d for d in pending_ads if d["status"] == "Draft"]),
        "in_production":   len([d for d in pending_ads if d["status"] == "En Producción"]),
        "errors":          len(report["errors"]),
    }
    s = report["summary"]
    print(f"\n   ✔  {config['name']} — Activos:{s['total_active']} | Pausados:{s['paused_today']} | Rotar:{s['rotation_alerts']} | Cola:{s['drafts']+s['in_production']}")
    return report


# ─────────────────────────────────────────────────────────────────────────────
#  NARRATIVE MEMO
# ─────────────────────────────────────────────────────────────────────────────

def build_narrative_memo(reports: list, run_date: str, dry_run: bool) -> str:
    lines = [
        f"# SPEKGEN MORNING INTELLIGENCE — {run_date}",
        "> ⚠️ DRY RUN — sin cambios reales" if dry_run else "",
        "",
        "## SITUATION — Estado general",
        "",
    ]
    for r in reports:
        s  = r.get("summary", {})
        cu = r.get("clickup_tasks", {})
        lines.append(f"### {r['client']}")
        if r.get("fatal_error"):
            lines.append(f"- **ERROR FATAL:** {r['fatal_error']}")
        else:
            cola = s.get("drafts", 0) + s.get("in_production", 0)
            lines.append(f"- Ads activos: {s.get('total_active', 0)} | En cola: {cola}")
            lines.append(f"- ClickUp: {cu.get('open', 0)} tareas abiertas | {cu.get('overdue', 0)} vencidas")
        lines.append("")

    lines += ["## COMPLICATIONS — Problemas detectados", ""]
    has_issues = False
    for r in reports:
        issues = []
        for p in r.get("ads_paused", []):
            issues.append(
                f"- **CPA ALTO:** `{p['ad_id']}` ({p['product']}) — "
                f"CPA ${p['cpa']} > límite ${p['threshold']} — PAUSADO"
            )
        for a in r.get("rotation_alerts", []):
            issues.append(
                f"- **ROTACIÓN:** `{a['ad_id']}` ({a['product']}) — "
                f"{a['days_active']} días activo"
            )
        for t in r.get("clickup_tasks", {}).get("tasks", []):
            if t["overdue"]:
                issues.append(f"- **TAREA VENCIDA:** {t['name']} ({r['client']})")
        for e in r.get("errors", []):
            issues.append(f"- **ERROR:** {e}")
        if issues:
            has_issues = True
            lines.append(f"### {r['client']}")
            lines.extend(issues)
            lines.append("")
    if not has_issues:
        lines += ["Sin complicaciones detectadas.", ""]

    lines += ["## RESOLUTION — Acciones tomadas y recomendadas", ""]
    for r in reports:
        actions = []
        for p in r.get("ads_paused", []):
            actions.append(f"- ✅ Pausado `{p['ad_id']}` en Meta {'(DRY RUN)' if dry_run else ''}")
        for a in r.get("rotation_alerts", []):
            actions.append(f"- \U0001f504 Crear nuevo creativo para `{a['ad_id']}` ({a['product']})")
        for t in r.get("clickup_tasks", {}).get("tasks", []):
            if t["overdue"]:
                actions.append(f"- ⚡ Resolver tarea vencida: {t['name']}")
        if actions:
            lines.append(f"### {r['client']}")
            lines.extend(actions)
            lines.append("")

    lines += [
        "---",
        f"*Generado automáticamente por spekgen_morning_intelligence.py — {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
    ]
    return "\n".join(line for line in lines if line is not None)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="SPEKGEN Morning Intelligence v2.0")
    parser.add_argument("--dry-run", action="store_true", help="Sin cambios en Meta ni AD LOG")
    parser.add_argument("--client", choices=list(CLIENTS.keys()) + ["ALL"], default="ALL")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  \U0001f305  SPEKGEN MORNING INTELLIGENCE v2.0")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if args.dry_run:
        print(f"  \U0001f9ea  MODO DRY RUN — sin cambios reales")
    print(f"{'='*60}")

    # Cargar .env local si existe (credenciales, nunca commitear)
    local_env_path = Path(__file__).parent / ".env"
    if local_env_path.exists():
        local_env = load_env_from_text(local_env_path.read_text(encoding="utf-8"))
        for k, v in local_env.items():
            if k not in os.environ:
                os.environ[k] = v

    google_token  = get_google_token()
    clickup_token = os.environ.get("CLICKUP_TOKEN", "")
    if not clickup_token:
        print("  ⚠️  CLICKUP_TOKEN no encontrado — tareas de ClickUp no disponibles")

    clients_to_run = {
        k: v for k, v in CLIENTS.items()
        if args.client == "ALL" or k == args.client
    }

    all_reports = []
    for client_key, config in clients_to_run.items():
        try:
            report = run_client(client_key, config, google_token, clickup_token, dry_run=args.dry_run)
            all_reports.append(report)
        except Exception as e:
            print(f"\n✗ Error fatal en {config['name']}: {e}")
            all_reports.append({
                "client":          config["name"],
                "client_key":      client_key,
                "clickup_list_id": config["clickup_list_id"],
                "fatal_error":     str(e),
                "ads_paused":      [],
                "rotation_alerts": [],
                "draft_queue":     [],
                "clickup_tasks":   {"open": 0, "overdue": 0, "tasks": []},
                "errors":          [str(e)],
                "summary":         {},
            })

    run_date = datetime.now().strftime("%Y-%m-%d")
    memo     = build_narrative_memo(all_reports, run_date, args.dry_run)

    print(f"\n{'='*60}")
    print(memo)
    print(f"{'='*60}")

    # Crear Google Doc en Drive
    if REPORTS_FOLDER_DRIVE_ID and "PLACEHOLDER" not in REPORTS_FOLDER_DRIVE_ID:
        doc_title = f"SPEKGEN Morning Intelligence — {run_date}"
        print(f"\n  ↑ Creando Google Doc en Drive...")
        doc_id = create_google_doc(doc_title, memo, REPORTS_FOLDER_DRIVE_ID, google_token)
        if doc_id:
            print(f"  ✔  Google Doc: https://docs.google.com/document/d/{doc_id}")
        else:
            print(f"  ✗  No se pudo crear el Google Doc")

    output = {
        "run_date":  run_date,
        "run_time":  datetime.now().strftime("%H:%M"),
        "dry_run":   args.dry_run,
        "memo":      memo,
        "reports":   all_reports,
    }
    print("\n__SPEKGEN_INTELLIGENCE_OUTPUT__")
    print(json.dumps(output, ensure_ascii=False, indent=2, default=str))
    print("__END_INTELLIGENCE_OUTPUT__")


if __name__ == "__main__":
    main()
